# Developed by Ahmed Abduljalil aka github username 0maido
# Project Name : Attendance Processor
# Date : 2025-02-25


from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import shutil
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import os

class AttendanceApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Attendance Processor")
        self.root.geometry("650x650")
        
        # Initialize variables
        self.main_file_path = None
        self.attendance_path = None
        self.total_attendance = 0
        self.in_main_count = 0
        self.not_in_main_count = 0
        
        # Configure styles
        self.style = ttk.Style()
        self.style.configure("TButton", padding=6)
        self.style.configure("Accent.TButton", foreground="white", background="#3498db")

        # Create main container
        main_frame = ttk.Frame(root, padding=20)
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Header Section
        header_frame = ttk.Frame(main_frame)
        header_frame.pack(pady=10, fill=tk.X)
        ttk.Label(header_frame, text="Attendance Processor", font=("Arial", 14, "bold")).pack()
        ttk.Label(header_frame, text="Developed by Ahmed Abduljalil", font=("Arial", 10)).pack()

        # File Selection Section
        file_frame = ttk.LabelFrame(main_frame, text="File Selection", padding=15)
        file_frame.pack(pady=10, fill=tk.X)

        # Main File Selection
        self.main_file_btn = ttk.Button(
            file_frame, 
            text="Select Main File", 
            command=self.select_main_file,
            style="Accent.TButton"
        )
        self.main_file_btn.pack(pady=5, anchor=tk.W)
        self.main_file_label = ttk.Label(file_frame, text="No main file selected")
        self.main_file_label.pack(pady=5, anchor=tk.W)

        # Attendance File Selection
        self.att_file_btn = ttk.Button(
            file_frame, 
            text="Select Attendance File", 
            command=self.select_attendance_file,
            style="Accent.TButton"
        )
        self.att_file_btn.pack(pady=5, anchor=tk.W)
        self.att_file_label = ttk.Label(file_frame, text="No attendance file selected")
        self.att_file_label.pack(pady=5, anchor=tk.W)

        # Range Settings
        range_frame = ttk.LabelFrame(main_frame, text="Processing Ranges", padding=15)
        range_frame.pack(pady=10, fill=tk.X)

        ttk.Label(range_frame, text="Attendance File Rows (e.g., 3-38):").grid(row=0, column=0, sticky=tk.W)
        self.att_row_entry = ttk.Entry(range_frame, width=15)
        self.att_row_entry.grid(row=0, column=1, padx=5)

        ttk.Label(range_frame, text="Attendance File Column:").grid(row=0, column=2, padx=5, sticky=tk.W)
        self.att_col_entry = ttk.Entry(range_frame, width=5)
        self.att_col_entry.grid(row=0, column=3)
        self.att_col_entry.insert(0, "C")

        ttk.Label(range_frame, text="Main File Rows (e.g., 3-38):").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.main_row_entry = ttk.Entry(range_frame, width=15)
        self.main_row_entry.grid(row=1, column=1, padx=5)

        ttk.Label(range_frame, text="Main File Column:").grid(row=1, column=2, padx=5, sticky=tk.W)
        self.main_col_entry = ttk.Entry(range_frame, width=5)
        self.main_col_entry.grid(row=1, column=3)
        self.main_col_entry.insert(0, "F")

        # Results Section
        results_frame = ttk.LabelFrame(main_frame, text="Results", padding=15)
        results_frame.pack(pady=10, fill=tk.X)

        self.results_labels = {
            'total': ttk.Label(results_frame, text="Total Students in the Attendence File: 0"),
            'present': ttk.Label(results_frame, text="Present: 0"),
            'absent': ttk.Label(results_frame, text="Absent: 0"),
            'wrong' : ttk.Label(results_frame, text="Students Who did it Wrong : 0")
        }
        self.results_labels['total'].grid(row=0, column=0, padx=10)
        self.results_labels['present'].grid(row=0, column=1, padx=10)
        self.results_labels['absent'].grid(row=0, column=2, padx=10)
        self.results_labels['wrong'].grid(row=0, column=3, padx=10)

        # Action Buttons
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(pady=10)

        self.process_btn = ttk.Button(
            btn_frame, 
            text="Process Attendance", 
            command=self.process_attendance, 
            state=tk.DISABLED
        )
        self.process_btn.pack(side=tk.LEFT, padx=5)

        self.export_btn = ttk.Button(
            btn_frame, 
            text="Export Results", 
            command=self.export_results, 
            state=tk.DISABLED
        )
        self.export_btn.pack(side=tk.LEFT, padx=5)

    def select_main_file(self):
        self.main_file_path = filedialog.askopenfilename(
            title="Select Main File",
            filetypes=[("Excel files", "*.xlsx;*.xls")]
        )
        if self.main_file_path:
            self.main_file_label.config(text=os.path.basename(self.main_file_path))
            self.check_files_selected()

    def select_attendance_file(self):
        self.attendance_path = filedialog.askopenfilename(
            title="Select Attendance File",
            filetypes=[("Excel files", "*.xlsx;*.xls")]
        )
        if self.attendance_path:
            self.att_file_label.config(text=os.path.basename(self.attendance_path))
            self.check_files_selected()

    def check_files_selected(self):
        if self.main_file_path and self.attendance_path:
            self.process_btn.config(state=tk.NORMAL)
        else:
            self.process_btn.config(state=tk.DISABLED)

    def validate_range(self, range_str):
        try:
            start, end = map(int, range_str.split('-'))
            return start, end
        except:
            messagebox.showerror("Error", "Invalid range format! Use start-end (e.g., 3-38)")
            return None

    def process_attendance(self):
        try:
            # Get processing ranges
            att_rows = self.validate_range(self.att_row_entry.get())
            main_rows = self.validate_range(self.main_row_entry.get())
            att_col = ord(self.att_col_entry.get().upper()) - 64  # Convert Column letter to number
            main_col = ord(self.main_col_entry.get().upper()) - 64

            # Create working copy of main file
            copy_path = os.path.join(os.path.dirname(self.main_file_path), "temp_copy.xlsx")
            shutil.copy(self.main_file_path, copy_path)

            # Load workbooks
            attendance_wb = load_workbook(self.attendance_path)
            main_wb = load_workbook(copy_path)

            # Get sheets
            att_sheet = attendance_wb.active
            main_sheet = main_wb.active

            # Collect student IDs
            present_ids = set()
            self.total_attendance = 0
            
            # Process attendance file
            for row in range(att_rows[0], att_rows[1] + 1):
                student_id = att_sheet.cell(row=row, column=att_col).value
                if student_id:
                    present_ids.add(str(student_id).strip())
                    self.total_attendance += 1

            # Process main file and mark attendance
            self.in_main_count = 0
            self.not_in_main_count = 0
            green_fill = PatternFill(start_color="00FF00", fill_type="solid")
            red_fill = PatternFill(start_color="FF0000", fill_type="solid")

            for row in range(main_rows[0], main_rows[1] + 1):
                student_id = str(main_sheet.cell(row=row, column=main_col).value).strip()
                status_cell = main_sheet.cell(row=row, column=7)  # Column G

                if student_id in present_ids:
                    status_cell.value = "P"
                    status_cell.fill = green_fill
                    self.in_main_count += 1
                else:
                    status_cell.value = "A"
                    status_cell.fill = red_fill
                    self.not_in_main_count += 1

            # Update results display
            self.results_labels['total'].config(text=f"Total Students in the Attendence File: {self.total_attendance}")
            self.results_labels['present'].config(text=f"Present: {self.in_main_count}")
            self.results_labels['absent'].config(text=f"Absent: {self.not_in_main_count}")
            self.results_labels['wrong'].config(text=f"Students Who did it wrong: {self.total_attendance-self.in_main_count}")

            # Save processed file
            main_wb.save(copy_path)
            self.export_btn.config(state=tk.NORMAL)
            messagebox.showinfo("Success", "Attendance processed successfully!\nClick 'Export Results' to save.")

        except Exception as e:
            messagebox.showerror("Error", f"Processing failed: {str(e)}")

    def export_results(self):
        output_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="Save Processed File As"
        )
        if output_path:
            try:
                # Copy temporary file to final destination
                temp_path = os.path.join(os.path.dirname(self.main_file_path), "temp_copy.xlsx")
                shutil.copy(temp_path, output_path)
                
                # Add statistics sheet
                wb = load_workbook(output_path)
                stats_sheet = wb.create_sheet("Statistics")
                
                # Add statistics data
                stats_sheet.append(["Metric", "Value"])
                stats_sheet.append(["Total Students in the Attendence File", self.total_attendance])
                stats_sheet.append(["Present", self.in_main_count])
                stats_sheet.append(["Absent", self.not_in_main_count])
                
                wb.save(output_path)
                messagebox.showinfo("Success", f"File saved successfully to:\n{output_path}")
                
                # Clean up temporary file
                os.remove(temp_path)
                
            except Exception as e:
                messagebox.showerror("Export Error", f"Failed to save file: {str(e)}")

if __name__ == "__main__":
    root = tk.Tk()
    app = AttendanceApp(root)
    root.mainloop()
    